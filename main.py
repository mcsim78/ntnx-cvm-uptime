import getpass
import json
import concurrent.futures
import os
import threading
from time import sleep
import openpyxl
from loguru import logger
import paramiko
import ipaddress
import argparse

SSH_PORT = 22
SSH_USERNAME = 'nutanix'
SSH_PASSWORD = ''
SSH_TIMEOUT = 5  # in seconds

OUTPUT_DIR = 'output/'  # output directory for logs and result files
INPUT_FILE = 'prism_ips.txt'

GET_PART_ROWS = 3  # part of rows to get from sorted dictionary, where 3 means 1/3 of rows, 4 means 1/4 of rows, etc.
SAVE_TO_XLSX = True  # save results to xlsx file
# For command execution
RUN_MODE = 'uptime'
EXEC_COMMAND = ''

# For multithreading
AVAIL_DICT = {}
UPTIME_DICT = {}
ERRORS_DICT = {}
MAX_WORKERS = 4  # not more than 10 threads!!!
TEMP_CREDS = {}
AVAIL_DICT_LOCK = threading.Lock()
UPTIME_DICT_LOCK = threading.Lock()
ERRORS_DICT_LOCK = threading.Lock()
TEMP_CREDS_LOCK = threading.Lock()
INPUT_EVENT = threading.Event()
LOCK_TIMEOUT = 10  # in seconds. Max time to wait for lock release
MAX_AUTH_ATTEMPTS = 2  # max attempts to authenticate to server

logger.add(f'{OUTPUT_DIR}logs/main.log', format='{time:DD-MM-YY HH:mm:ss} - {level} - {message}',
           level='INFO', rotation='1 week', compression='zip')


class ResultReturn:
    """
    For returning results of operations
    """

    def __init__(self, success=False, data=None, error=None):
        self.success = success
        self.data = data
        self.error = error


def attempt_to_auth_on_cvm(ssh_client: paramiko.SSHClient, prism_address: str) -> ResultReturn:
    result = ResultReturn()
    logger.warning(f'Authentication failed for for cluster "{prism_address}"')
    INPUT_EVENT.set()
    sleep(3)  # wait for other threads to stop printing to console
    temp_ssh_user = input(f'\nEnter username for cluster "{prism_address}"\n(Leave empty and press Enter to skip): ')
    temp_ssh_password = ''
    if temp_ssh_user:
        temp_ssh_password = getpass.getpass(f'Enter password for user "{temp_ssh_user}": ')
    INPUT_EVENT.clear()  # Clear event to allow other threads to continue
    if not temp_ssh_user or not temp_ssh_password:
        logger.warning(f'Skip for cluster "{prism_address}"')
        result.error = f'Username or password is empty for cluster "{prism_address}"'
        return result
    try:
        ssh_client.connect(hostname=prism_address, port=SSH_PORT, username=temp_ssh_user,
                           password=temp_ssh_password, timeout=SSH_TIMEOUT)
        logger.info(f'Authentication success for "{prism_address}"')
        sleep(0.5)
    except paramiko.AuthenticationException:
        logger.error(f'Authentication failed for "{prism_address}"')
        result.error = f'Authentication failed "{prism_address}"'
        return result
    except Exception as e:
        logger.error(f'Exception: {e}')
        result.error = f'Exception: {e}'
    # If no errors
    TEMP_CREDS_LOCK.acquire()
    try:
        TEMP_CREDS[prism_address] = {'username': temp_ssh_user, 'password': temp_ssh_password}
        result.success = True
        result.data = ssh_client
    finally:
        TEMP_CREDS_LOCK.release()

    return result


def get_svm_ips_from_server(server: str) -> ResultReturn:
    """
    Get SVM IPs from server. Use standard Nutanix command: /usr/local/nutanix/cluster/bin/svmips
    :param server: server IP or hostname
    :return: ResultReturn object
    """
    result = ResultReturn()
    svm_ips = []
    cmd_svm_ips = '/usr/local/nutanix/cluster/bin/svmips'
    auth_attempts = 1

    logger.info(f'+ Getting SVM IPs for server: {server}')
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh_client.connect(hostname=server, port=SSH_PORT, username=SSH_USERNAME,
                           password=SSH_PASSWORD, timeout=SSH_TIMEOUT)
        # sleep(0.5)
    except paramiko.AuthenticationException:
        if auth_attempts == MAX_AUTH_ATTEMPTS:
            logger.error(f'Authentication failed for for cluster "{server}". Skipping...')
            return result
        else:
            auth_attempts += 1
    except OSError as excpt:
        logger.error(f'Error occurred for "{server}": {excpt}')
        result.error = f'Error occurred for "{server}": {excpt}'
        return result

    # Try another auth attempt if first failed
    if auth_attempts == MAX_AUTH_ATTEMPTS:
        result = attempt_to_auth_on_cvm(ssh_client=ssh_client, prism_address=server)
        if result.success:
            ssh_client = result.data
        else:
            return result
    # get svm ips
    stdin, stdout, stderr = ssh_client.exec_command(cmd_svm_ips)
    std_out = stdout.readlines()
    out_error = stderr.readlines()
    if not out_error and std_out:
        # get svm ips
        ips = std_out[0].strip()
        try:
            # Check if ip is valid
            ips = ips.split()
            for i in ips:
                ipaddress.IPv4Address(i)
                svm_ips.append(i)
        except ValueError:
            logger.error(f' - Error getting svm_ips: {ips}')
            result.error = f'Error getting svm_ips: {ips}'
            return result
        result.success = True
        result.data = svm_ips
    else:
        logger.warning(f' - std_out for server {server} is probably empty. Skipped...')
        result.error = f'std_out for server {server} is probably empty. Skipped...'
    ssh_client.close()
    # logger.info(f'  No errors')
    return result


def get_prism_addresses_from_file(filepath: str) -> []:
    """
    Load SVM IPs from file to list
    :param filepath:
    :return: list of SVM IPs
    """
    if not os.path.isfile(filepath):
        logger.error(f'File "{filepath}" not found')
        return []
    svm_ips = []
    with open(filepath, 'r') as f:
        for line in f:
            if line.strip() == '' or line.strip().startswith('#'):
                continue
            svm_ips.append(line.strip())
    return svm_ips


def get_average_from_dict(svm_dict: {}, key: str, cut_off=False) -> int:
    """
    Get average value from dictionary
    """
    # Sort dictionary by value
    sort = False if key == 'available' else True
    sliced_dict = sort_dict_by_key(dict_to_sort=svm_dict, key_name=key, reverse_sort=sort)
    if cut_off:
        sliced_dict = get_first_rows(sliced_dict, GET_PART_ROWS)
    if not sliced_dict:
        return 0
    summ = 0
    for svm in sliced_dict:
        summ += int(sliced_dict[svm][key])
    return int(summ / len(sliced_dict))


def get_tmp_creds(prism_address):
    """
    Get user's entered credentials for ssh if they exist, otherwise return defaults
    :param prism_address: Address (ip or fqdn) of Prism to get credentials for
    :return: Tuple with username and password
    """
    try:
        tmp_ssh_user = TEMP_CREDS[prism_address]['username']
        tmp_ssh_password = TEMP_CREDS[prism_address]['password']
    except KeyError:
        tmp_ssh_user = SSH_USERNAME
        tmp_ssh_password = SSH_PASSWORD
    return tmp_ssh_password, tmp_ssh_user


def get_all_ssh_available_mem(prism_address: str, svm_list: []) -> {}:
    """
    Get available memory for all SVMs
    :param prism_address: Address (ip or fqdn) of Prism
    :param svm_list: list of ip addresses for getting available memory
    :return: Dictionary with ip addresses as keys and available memory as values
    """
    cmd = "free | grep '^Mem:' | awk '{print $7}'"
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    svm_dict = {}
    # Try to get temporary credentials if they exist for this cluster, otherwise use default
    tmp_ssh_password, tmp_ssh_user = get_tmp_creds(prism_address)

    for ip in svm_list:
        while INPUT_EVENT.is_set():
            INPUT_EVENT.wait(LOCK_TIMEOUT)

        ssh_client.close()
        try:
            ssh_client.connect(hostname=ip, port=SSH_PORT, username=tmp_ssh_user,
                               password=tmp_ssh_password, timeout=SSH_TIMEOUT)
            # sleep(0.5)
        except Exception as e:
            svm_dict[ip] = {
                'available': 0,
                'error': f'Connection failed: {e}'
            }
            logger.error(f' - Connection for {ip} failed: {e}')
            continue
        logger.info(f'+ Getting available memory for {ip}')
        stdin, stdout, stderr = ssh_client.exec_command(cmd)
        std_out = stdout.readlines()
        out_error = stderr.readlines()
        if not out_error and std_out:
            available = std_out[0].strip()
            svm_dict[ip] = {
                'available': available,
                'error': 'none'
            }
        else:
            svm_dict[ip] = {
                'available': 0,
                'error': f'Error exec command "{cmd}"'
            }
            logger.error(f' - Error exec command "{cmd}" on {ip}: {out_error}')
    ssh_client.close()
    # logger.info(f'  No errors')
    save_to_json(svm_dict, f'{OUTPUT_DIR}avail_{prism_address}.json')
    # logger.info(
    #     f'+ Getting average available memory for {prism_address} and cutting off 1/{GET_PART_ROWS} part of rows')
    # Cut off unneeded part of dictionary
    # Summ all available memory and get average of left CVMs
    cut_off = False if len(svm_dict) <= GET_PART_ROWS else True
    average = get_average_from_dict(svm_dict, 'available', cut_off=cut_off)
    # Write down to dictionary with prism_address as key and available memory as value
    AVAIL_DICT_LOCK.acquire()
    try:
        AVAIL_DICT[prism_address] = {
            'available': average,
            'hosts_count': len(svm_dict),
        }
    finally:
        AVAIL_DICT_LOCK.release()
    return AVAIL_DICT


def get_log_errors(prism_address: str, svm_list: []) -> {}:
    if not EXEC_COMMAND:
        # Default command for getting errors from hades.out
        cmd = "sudo grep -i 'Failed to get CVM id' ~/data/logs/hades.out | tail -n 2"
    else:
        cmd = EXEC_COMMAND
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    errors_dict = {}
    # Try to get temporary credentials if they exist for this cluster, otherwise use default
    tmp_ssh_password, tmp_ssh_user = get_tmp_creds(prism_address)

    for ip in svm_list:
        while INPUT_EVENT.is_set():
            INPUT_EVENT.wait(LOCK_TIMEOUT)

        ssh_client.close()
        try:
            ssh_client.connect(hostname=ip, port=SSH_PORT, username=tmp_ssh_user,
                               password=tmp_ssh_password, timeout=SSH_TIMEOUT)
            # sleep(0.5)
        except Exception as e:
            errors_dict[ip] = {
                'logs': '',
                'error': f'Connection failed: {e}'
            }
            logger.error(f' - Connection for {ip} failed: {e}')
            continue
        logger.info(f'+ Getting errors in logs for {ip}')
        stdin, stdout, stderr = ssh_client.exec_command(cmd)
        std_out = stdout.readlines()
        out_error = stderr.readlines()
        if not out_error and std_out:
            logs = std_out
            errors_dict[ip] = {
                'logs': logs,
                'error': 'none'
            }
        elif not out_error and not std_out:
            # No errors in logs, skip
            continue
        else:
            errors_dict[ip] = {
                'logs': '',
                'error': f'Error exec command "{cmd}"'
            }
            logger.error(f' - Error exec command "{cmd}" on {ip}: {out_error}')
    ssh_client.close()
    # logger.info(f'  No errors')
    save_to_json(errors_dict, f'{OUTPUT_DIR}errors_{prism_address}.json')
    # Write down to dictionary with prism_address as key and available memory as value
    if errors_dict:
        ERRORS_DICT_LOCK.acquire()
        ERRORS_DICT[prism_address] = errors_dict
        ERRORS_DICT_LOCK.release()
    return ERRORS_DICT


def get_all_ssh_uptime(prism_address: str, svm_list: []) -> {}:
    """
    Get uptime for all SVMs
    :param prism_address:
    :param svm_list: List of ip addresses for getting uptime
    :return: Dictionary with ip addresses as keys and uptime as values
    """
    # cmd = "uptime | grep days | awk '{print $3}'"
    cmd = "cat /proc/uptime | awk '{print $1}'"  # uptime in seconds
    ssh_client = paramiko.SSHClient()
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    svm_dict = {}
    # Try to get temporary credentials if they exist for this cluster, otherwise use default
    tmp_ssh_password, tmp_ssh_user = get_tmp_creds(prism_address)
    for ip in svm_list:
        while INPUT_EVENT.is_set():
            INPUT_EVENT.wait(LOCK_TIMEOUT)

        ssh_client.close()
        try:
            ssh_client.connect(hostname=ip, port=SSH_PORT, username=tmp_ssh_user,
                               password=tmp_ssh_password, timeout=SSH_TIMEOUT)
            # sleep(0.5)
        except Exception as e:
            svm_dict[ip] = {
                'uptime': 0,
                'error': f'Connection failed: {e}'
            }
            logger.error(f' - Connection for {ip} failed: {e}')
            continue
        logger.info(f'+ Getting uptime for {ip}')
        stdin, stdout, stderr = ssh_client.exec_command(cmd)
        std_out = stdout.readlines()
        out_error = stderr.readlines()
        if not out_error and std_out:
            up_sec = float(std_out[0].strip())
            if up_sec > 86400:
                uptime = round(up_sec / 86400)
                svm_dict[ip] = {
                    'uptime': uptime,
                    'error': 'none'
                }
            else:
                svm_dict[ip] = {
                    'uptime': 0,
                    'error': f'Uptime less than 1 day'
                }
            # uptime = std_out[0].strip()
        else:
            svm_dict[ip] = {
                'uptime': 0,
                'error': f'Error exec command "{cmd}"'
            }
            logger.error(f' - Error exec command "{cmd}" for {ip}')
    ssh_client.close()
    # logger.info(f'  No errors')
    save_to_json(svm_dict, f'{OUTPUT_DIR}uptime_{prism_address}.json')
    # logger.info(f'+ Getting average uptime for {prism_address} and cutting off 1/{GET_PART_ROWS} part of rows')
    # 1.Cut off unneeded part of dictionary
    # 2.Summ all uptime and get average of left CVMs
    cut_off = False if len(svm_dict) <= GET_PART_ROWS else True
    average = get_average_from_dict(svm_dict, 'uptime', cut_off=cut_off)
    # Write down to dictionary with prism_address as key and uptime as value
    UPTIME_DICT_LOCK.acquire()
    try:
        UPTIME_DICT[prism_address] = {
            'uptime': average,
            'hosts_count': len(svm_dict),
        }
    finally:
        UPTIME_DICT_LOCK.release()
    return UPTIME_DICT


def sort_dict_by_key(dict_to_sort: {}, key_name: str, reverse_sort: True) -> {}:
    """
    Sort dict by key name
    :param reverse_sort: If True - sort in reverse order, else - sort in normal order
    :param key_name: Key name to sort by
    :param dict_to_sort: Dictionary to sort
    :return: sorted dict
    """
    srt_dict = dict(sorted(dict_to_sort.items(), key=lambda item: (float(item[1][key_name])), reverse=reverse_sort))
    return srt_dict


def save_to_json(dict_to_save: {}, filepath: str):
    """
    Save dictionary to json file
    :param dict_to_save: What to save
    :param filepath: Where to save
    :return: Nothing
    """
    with open(filepath, 'w') as f:
        json.dump(dict_to_save, f, indent=4)
        # logger.info(f'+ Result saved to {filepath}')


def save_to_xslxs(dict_to_save: {}, filepath: str):
    """
    Save dictionary to xlsx file
    :param dict_to_save: What to save
    :param filepath: Where to save
    :return: Nothing
    """
    wb_obj = openpyxl.Workbook()
    sheet_obj = wb_obj.active
    sheet_obj.cell(row=2, column=1).value = '#'
    sheet_obj.cell(row=2, column=2).value = 'Cluster'
    sheet_obj.cell(row=2, column=3).value = 'Uptime, days'
    sheet_obj.cell(row=2, column=4).value = 'Hosts'
    sheet_obj.cell(row=2, column=5).value = 'Cluster'
    sheet_obj.cell(row=2, column=6).value = 'Free memory, GB'
    sheet_obj.cell(row=2, column=7).value = 'Hosts'
    row = 3
    for i, (k, v) in enumerate(dict_to_save.items()):
        sheet_obj.cell(row=row, column=1).value = k
        sheet_obj.cell(row=row, column=2).value = v['cluster_uptime']
        sheet_obj.cell(row=row, column=3).value = v['uptime']
        sheet_obj.cell(row=row, column=4).value = v['cluster_up_hosts']
        sheet_obj.cell(row=row, column=5).value = v['cluster_avail']
        sheet_obj.cell(row=row, column=6).value = v['available']
        sheet_obj.cell(row=row, column=7).value = v['cluster_avail_hosts']
        row += 1
    # Calculate and set columns width
    columns_letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for letter in columns_letter:
        cur_max_width = 0
        for cell in sheet_obj[letter]:
            if cell.value is None:
                continue
            new_width = len(str(cell.value))
            if new_width > cur_max_width:
                cur_max_width = new_width
        col_dim = sheet_obj.column_dimensions[letter]
        col_dim.width = cur_max_width + 1
    # Set headers
    sheet_obj.cell(row=1, column=1).value = 'Uptime (average)'
    sheet_obj.cell(row=1, column=5).value = 'Free (average)'
    sheet_obj.merge_cells('A1:D1')
    sheet_obj.merge_cells('E1:G1')
    wb_obj.save(filepath)


def get_first_rows(data_dict: {}, divider: int) -> {}:
    """
    Get first rows from dictionary
    :param data_dict: Source dictionary
    :param divider: What part of dictionary to get. For example, if divider=2, then get first 1/2 (half) of dictionary
    :return: Dictionary with filtered rows
    """
    rows_count = divmod(len(data_dict), divider)[0]  # Getting an integer from division
    rows_count = 1 if rows_count == 0 else rows_count
    sliced_dict = {k: v for i, (k, v) in enumerate(data_dict.items()) if i < rows_count}
    return sliced_dict


def print_dicts(uptime_dict, avail_dict):
    new_dict = {}
    # Print table header
    print('===== Collected data =====')
    print("{:<55}{:<30}".format("Uptime (average)", "Free (average)"))

    # Get keys from first dictionaries
    keys = list(uptime_dict.keys())

    # Iterate over keys and print values from both dictionaries. Second dictionary left origin ordered.
    for i, key in enumerate(keys):
        # Get values from both dictionaries
        if i + 1 > len(avail_dict.keys()):
            value2 = {
                'available': 0,
                'hosts_count': 0,
            }
            key2 = '---'
        else:
            key2 = list(avail_dict.keys())[i]
            value2 = avail_dict[key2]

        value1 = uptime_dict[key]

        # Print values
        avail_gb = round(float(value2["available"]) / 1024 / 1024, 2)
        print("{:>2d}. {:<42} {:>3} up | {:<30} {:>5} GB".format(i + 1, key, value1["uptime"], key2, avail_gb))
        # Save values to new dictionary for xlsx file
        new_dict[i + 1] = {
            'cluster_uptime': key,
            'cluster_up_hosts': value1['hosts_count'],
            'uptime': value1['uptime'],
            'cluster_avail': key2,
            'cluster_avail_hosts': value2['hosts_count'],
            'available': avail_gb,
        }
    print('========================\n')
    return new_dict


def print_dict_log_errors(dict_with_errors: dict):
    """
    Print dictionary with errors
    :param dict_with_errors: Dictionary with errors
    :return: Nothing
    """
    if dict_with_errors:
        print('===== Errors =====')
        for cluster, cvms in dict_with_errors.items():
            if not cvms:
                continue
            print(f'{cluster}:')
            for cvm, logs in cvms.items():
                print(f'\tCVM {cvm}:')
                for error in logs['logs']:
                    print(f'\t\t{error.strip()}')


def start():
    prism_addresses = get_prism_addresses_from_file(INPUT_FILE)
    if not prism_addresses:
        logger.error(f'No prism addresses found in {INPUT_FILE}')
        return
    # Running in threads to speed up
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = []
        for prism_address in prism_addresses:
            future = executor.submit(process_prism, prism_address)
            futures.append(future)
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logger.error(f'Exception: {e}')


def process_prism(prism_address):
    while INPUT_EVENT.is_set():
        INPUT_EVENT.wait(LOCK_TIMEOUT)
    svm_ips = get_svm_ips_from_server(prism_address)
    if svm_ips.success:
        svm_ips = svm_ips.data
        if RUN_MODE == 'uptime':
            # Get available memory
            get_all_ssh_available_mem(prism_address, svm_ips)
            # Get uptime
            get_all_ssh_uptime(prism_address, svm_ips)
        elif RUN_MODE == 'check_errors':
            # Get errors in logs
            get_log_errors(prism_address, svm_ips)
    else:
        logger.error(f'Error get SVM IPs from {prism_address}: {svm_ips.error}')


def main():
    global SSH_USERNAME, SSH_PASSWORD
    if not SSH_USERNAME:
        logger.warning('No username provided. Will be asked during script execution.')
        SSH_USERNAME = input(f'Enter SSH username that can be used on the most clusters: ')
        if not SSH_USERNAME:
            logger.error('No username provided')
            return False
    if not SSH_PASSWORD:
        logger.warning('No password provided. Will be asked during script execution.')
        SSH_PASSWORD = getpass.getpass(
            prompt=f'Enter SSH password for user "{SSH_USERNAME}" that can be used on the most clusters: '
        )
        if not SSH_PASSWORD:
            logger.error('No password provided')
            return False

    # Start collecting data
    try:
        logger.info('===== Start collecting data =====')
        start()
    except Exception as e:
        logger.error(f'Error: {e}')
        raise e
    logger.info('===== End collecting data =====\n')
    # Sort and print dicts
    if AVAIL_DICT and UPTIME_DICT:
        # Get sorted dicts
        sorted_avail = sort_dict_by_key(AVAIL_DICT, key_name='available', reverse_sort=False)
        sorted_uptime = sort_dict_by_key(UPTIME_DICT, key_name='uptime', reverse_sort=True)
        # Print sorted dicts
        merged_dict = print_dicts(sorted_uptime, sorted_avail)
        save_to_json(merged_dict, f'{OUTPUT_DIR}00_clusters_stat.json')
        print(f'Data saved to {OUTPUT_DIR}00_clusters_stat.json\n')
        if SAVE_TO_XLSX:
            filename = f'{OUTPUT_DIR}00_clusters_stat.xlsx'
            save_to_xslxs(merged_dict, filename)
            print(f'Data saved to {filename}\n')
    if ERRORS_DICT:
        print_dict_log_errors(ERRORS_DICT)
        save_to_json(ERRORS_DICT, f'{OUTPUT_DIR}00_clusters_errors.json')
        print(f'Data saved to {OUTPUT_DIR}00_clusters_errors.json\n')
    else:
        print('No logs errors found')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Сбор статистики и ошибок по кластерам Prism. Запуск без параметров '
                                                 'выполнится в режиме uptime.')
    parser.add_argument('-m', '--mode',
                        help='Режим запуска: '
                             'uptime - сбор uptime и free memory; '
                             'check_errors - проверка наличия ошибок в логах. Если выбран "check_errors" режим, '
                             'то параметр "-cmd" обязателен.',
                        choices=['uptime', 'check_errors'], default='uptime',
                        )
    parser.add_argument('-cmd', '--command', help='Строка с командой для проверки логов', type=str)
    parser.add_argument('-w', '--workers', help='(Необязательный) Количество потоков (не более 10)',
                        type=int, required=False)
    # parser.add_argument('-i', '--input', help='Файл с адресами Prism', default='prism_ips.txt', required=False)
    args = parser.parse_args()
    if args.mode:
        RUN_MODE = args.mode
    if RUN_MODE == 'check_errors' and not args.command:
        logger.error('No command provided')
        exit(1)
    if args.command:
        EXEC_COMMAND = args.command
    if args.workers:
        MAX_WORKERS = args.workers
    # if args.input:
    #     INPUT_FILE = args.input

    try:
        main()
    except Exception as e:
        logger.error(f'Error: {e}')
        raise e
